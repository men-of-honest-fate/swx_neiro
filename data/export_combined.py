"""
Объединяет enriched-каталоги циклов 23–25 в один XLSX.
Строки с отрицательными T_delta_flare или T_delta_SPE подсвечиваются цветом:
  - красный  (FFC7CE) — T_delta_SPE < 0
  - жёлтый   (FFEB9C) — T_delta_flare < 0
  - оранжевый (FFCC99) — оба < 0
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import os

RESULTS_DIR = os.path.join(os.path.dirname(__file__), '..', 'results')
OUTPUT = os.path.join(RESULTS_DIR, 'spe_catalog_combined.xlsx')

PATHS = {
    23: os.path.join(RESULTS_DIR, 'spe_catalog_23cycle_full_enriched.csv'),
    24: os.path.join(RESULTS_DIR, 'spe_catalog_24cycle_full_enriched.csv'),
    25: os.path.join(RESULTS_DIR, 'spe_catalog_25cycle_full_enriched.csv'),
}

FILL_RED    = PatternFill(fill_type='solid', fgColor='FFC7CE')  # T_delta_SPE < 0
FILL_YELLOW = PatternFill(fill_type='solid', fgColor='FFEB9C')  # T_delta_flare < 0
FILL_ORANGE = PatternFill(fill_type='solid', fgColor='FFCC99')  # оба < 0

# ---------- загрузка данных ----------
frames = []
for cycle, path in PATHS.items():
    df = pd.read_csv(path)
    df['Cycle'] = cycle  # перезаписать на случай если колонка уже есть
    frames.append(df)

combined = pd.concat(frames, ignore_index=True)

# Убрать технические колонки парсера PDF (не нужны в итоговой таблице)
drop_cols = ['Page_in_PDF']
combined.drop(columns=[c for c in drop_cols if c in combined.columns], inplace=True)

# Перенести Cycle в начало
cols = ['Cycle'] + [c for c in combined.columns if c != 'Cycle']
combined = combined[cols]

print(f'Всего строк: {len(combined)}')
print(f'  Цикл 23: {(combined.Cycle == 23).sum()}')
print(f'  Цикл 24: {(combined.Cycle == 24).sum()}')
print(f'  Цикл 25: {(combined.Cycle == 25).sum()}')

neg_spe   = combined['T_delta_SPE'].fillna(0) < 0
neg_flare = combined['T_delta_flare'].fillna(0) < 0
print(f'\nБракованных строк:')
print(f'  T_delta_SPE   < 0 : {neg_spe.sum()}')
print(f'  T_delta_flare < 0 : {neg_flare.sum()}')
print(f'  Оба           < 0 : {(neg_spe & neg_flare).sum()}')

# ---------- запись в Excel ----------
combined.to_excel(OUTPUT, index=False, sheet_name='SPE_catalog')
print(f'\nЗаписан файл: {OUTPUT}')

# ---------- окраска ----------
wb = load_workbook(OUTPUT)
ws = wb.active

# Заголовок — жирный
for cell in ws[1]:
    cell.font = Font(bold=True)

# Найти индексы нужных колонок
header = {cell.value: cell.column for cell in ws[1]}
col_spe   = header.get('T_delta_SPE')
col_flare = header.get('T_delta_flare')

n_colored = 0
for row_idx, (_, row) in enumerate(combined.iterrows(), start=2):
    bad_spe   = pd.notna(row['T_delta_SPE'])   and row['T_delta_SPE']   < 0
    bad_flare = pd.notna(row['T_delta_flare']) and row['T_delta_flare'] < 0

    if bad_spe and bad_flare:
        fill = FILL_ORANGE
    elif bad_spe:
        fill = FILL_RED
    elif bad_flare:
        fill = FILL_YELLOW
    else:
        continue

    # Красить всю строку
    for col_idx in range(1, ws.max_column + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill

    # Дополнительно — выделить ячейки с отрицательными значениями жирным
    if bad_spe and col_spe:
        ws.cell(row=row_idx, column=col_spe).font = Font(bold=True)
    if bad_flare and col_flare:
        ws.cell(row=row_idx, column=col_flare).font = Font(bold=True)

    n_colored += 1

# Автоширина первых колонок
for col in ws.iter_cols(min_row=1, max_row=1):
    col_letter = get_column_letter(col[0].column)
    max_len = max((len(str(cell.value or '')) for cell in ws[col_letter]), default=8)
    ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

wb.save(OUTPUT)
print(f'Подсвечено строк: {n_colored}')
print('\nЛегенда:')
print('  Красный  (FFC7CE) — T_delta_SPE < 0')
print('  Жёлтый   (FFEB9C) — T_delta_flare < 0')
print('  Оранжевый(FFCC99) — оба < 0')
